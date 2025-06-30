# excel_module.py
"""
openpyxlを利用し、Excelファイルを読み込むモジュール。

Classes:
    ExcelData: Excelファイルを読み込むクラス
    DataHandler: 読み込んだExcelDataを操作するクラス
"""

import openpyxl
import os
import sys

class ExcelData:
    """
    Excelファイルを読み込む。

    file_pathが指定されていない場合、load()メソッドを使用してファイルの読み込みが必要になります。
    
    Attributes:
        workbook (openpyxl.Workbook): 読み込んだExcelファイルのワークブックオブジェクト
        sheetnames (list): 読み込んだExcelファイルのシート名のリスト
        sheets (dict): シート名をキーとし、シートオブジェクトを値とする辞書
    
    Methods:

    """
    def __init__(self, file_path = None):
        self.workbook = None
        self.sheetnames = []
        self.sheets = {}
        if file_path:
            self.load(file_path)

    def __getitem__(self, sheet_name):
        """
        シート名でシートを取得
        """
        if self.workbook is None:
            raise ValueError("エクセルファイルが読み込んでいません。load()を使用してファイルを読み込んでください。")
        if sheet_name not in self.sheetnames:
            raise ValueError(f"シート名'{sheet_name}'が存在していません。")
        return self.sheets[sheet_name]

    def load(self, file_path):
        """
        Excelファイルを読み込む
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"指定されたファイル'{file_path}'が存在していません。")
        
        # Excelファイル
        self.workbook = openpyxl.load_workbook(file_path)
        # シート名
        self.sheetnames = self.workbook.sheetnames
        # シートの辞書
        self.sheets = {name: self.workbook[name] for name in self.sheetnames}


class DataHandler:
    """
    読み込んだExcelDataを操作するクラス。
    
    Attributes:
        data (ExcelData): ExcelDataのインスタンス。

    Methods:
        set_sheet(sheet_name): インスタンスのシート名を設定、set_sheet()を使用して、シート名を設定すると、以降のメソッドでsheet_nameを入力しなくても、そのシート名をデフォルトとして使用します。
        get_sheet(sheet_name): シートのデータを取得します。
        get_cell(cell, sheet_name): セルの値を取得します。
        get_row(row, sheet_name): 行のデータを取得します。
        get_column(column, sheet_name): 列のデータを取得します。
    """
    def __init__(self, excel_data, sheet_name = None):
        self.data = excel_data
        # ExcelDataのインスタンスであることを確認
        if not isinstance(self.data, ExcelData):
            raise ValueError("ExcelFileのインスタンスではありません。ExcelFileのインスタンスを渡してください。")
        self._current_sheet_name = None

        # 初期化時にシート名が指定された場合
        if sheet_name is not None:
            # シートの存在を確認
            if sheet_name not in self.data.sheetnames:
                raise ValueError(f"シート名 '{sheet_name}' が存在していません。")
            self.set_sheet(sheet_name)

    def _resolve_sheet_name(self, sheet_name):
        """
        使うシート名を決定する
        """
        # 関数実行時、シート名を指定しなかった場合
        if sheet_name is None:
            # 関数実行時、シート名が指定されていないばあい
            if self._current_sheet_name is None:
                # set_sheet() でシートめが設定されていない    
                raise ValueError(f"set_sheet('sheet_name')の実行、または {sys._getframe(1).f_code.co_name}(sheet_name = 'sheet_name')を指定してください")
            # set_sheet()　でシート名が設定されている場合
            return self._current_sheet_name
        
        # 関数実行時、シート名を指定した場合
        if sheet_name not in self.data.sheetnames:
            # 指定されたシート名が存在しない場合
            raise ValueError(f"シート名 '{sheet_name}' が存在していません。")
        
        # 指定されたシート名が存在する場合
        return sheet_name

    def _get_raw_cell_value(self, sheet, cell):
        """
        セルの生の値を取得する内部ユーティリティ関数。
        (MergedCellDataHandlerでオーバーライドされます)
        """
        return cell.value

    def set_sheet(self, sheet_name) -> None:
        """
        インスタンスのシート名を設定
        """
        if sheet_name not in self.data.sheetnames:
            raise ValueError(f"シート名 '{sheet_name}' が存在していません。")
        self._current_sheet_name = sheet_name

    def get_sheet(self, sheet_name = None):
        """
        シート名からシートデータを取得
        """
        solved_sheet_name = self._resolve_sheet_name(sheet_name)
        current_sheet = self.data[solved_sheet_name]
        sheet_data = []
        for row in current_sheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(self._get_raw_cell_value(current_sheet, cell))
            sheet_data.append(row_data)
        return sheet_data

    def get_cell(self, cell_coordinate, sheet_name=None):
        """
        セルの値を取得
        """
        solved_sheet_name = self._resolve_sheet_name(sheet_name)
        current_sheet = self.data[solved_sheet_name]
        cell_obj = current_sheet[cell_coordinate]
        return self._get_raw_cell_value(current_sheet, cell_obj)
    
    def get_row(self, row_index, sheet_name=None):
        """ 
        行のデータを取得
        """
        solved_sheet_name = self._resolve_sheet_name(sheet_name)
        current_sheet = self.data[solved_sheet_name]
        if row_index < 1:
            raise ValueError("行番号は1以上でなければなりません。")
        
        row_data = []
        for cell in current_sheet[row_index]:
            row_data.append(self._get_raw_cell_value(current_sheet, cell))
        return row_data
    
    def get_column(self, col_index, sheet_name=None):
        """
        列のデータを取得
        """
        solved_sheet_name = self._resolve_sheet_name(sheet_name)
        current_sheet = self.data[solved_sheet_name]
        if col_index < 1:
            raise ValueError("列番号は1以上でなければなりません。")
        
        column_data = []
        for cell_tuple in current_sheet.iter_cols(min_col=col_index, max_col=col_index):
            for cell in cell_tuple:
                column_data.append(self._get_raw_cell_value(current_sheet, cell))
        return column_data